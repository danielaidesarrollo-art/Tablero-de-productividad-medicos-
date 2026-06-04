import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px

# 1. Configuración de la página
st.set_page_config(page_title="Tablero Virrey Solís - Productividad", layout="wide")

# --- COLORES INSTITUCIONALES ---
COLOR_AZUL = "#004B87"
COLOR_VERDE = "#00A650"
COLOR_ALERTA = "#E74C3C" 

CORRECCIONES_MEDICOS = {
    "KARON": "CAROL",
    "KAROL": "CAROL",
    "JHON ERIC": "JHON ERIK",
}

# --- FUNCIONES DE COLOR MEJORADAS ---
def obtener_color_hex(cell):
    """Obtiene el color de la celda de la forma más precisa posible."""
    if not hasattr(cell, 'fill') or not cell.fill or not cell.fill.start_color:
        return "SIN_COLOR"
    
    color_val = str(cell.fill.start_color.index).upper()
    
    # Códigos internos de Excel para "Sin Relleno"
    if color_val in ['00000000', '0', '64', 'FFFFFFFF', '00FFFFFF', 'SYSTEM FOREGROUND']:
        return "SIN_COLOR"
        
    return color_val

def es_color_cafe(color_str):
    """Solo devuelve True si el color es EXPLÍCITAMENTE rojo, naranja oscuro o café."""
    if color_str == "SIN_COLOR": 
        return False
        
    # Códigos Hex comunes para Rojo, Naranja y Café en Excel
    rojos_cafes = ['FFFF0000', 'FFC00000', 'FFFFC000', 'FFE26B0A', 'FF974706', 'FFC65911', 'FFED7D31', 'FF806000']
    
    return any(c in color_str for c in rojos_cafes)

def es_color_azul(color_str):
    if color_str == "SIN_COLOR":
        return False
    azules = ['FF00B0F0', 'FF0070C0', 'FF4472C4', 'FF5B9BD5', 'FF1F497D', 'FFDEEBF7', 'FFBDD7EE', 'FF9BC2E6']
    return any(azul in color_str for azul in azules)

# 2. Procesamiento y Limpieza
@st.cache_data
def cargar_y_procesar_datos(archivo_subido):
    df = pd.read_excel(archivo_subido)
    df.columns = df.columns.astype(str).str.strip().str.upper()
    
    archivo_subido.seek(0)
    wb = openpyxl.load_workbook(archivo_subido, data_only=True)
    hoja = wb.active
    
    estados = []
    filas_validas = []
    colores_detectados = [] # Para depuración
    max_filas = len(df) + 1
    
    for idx, row in enumerate(hoja.iter_rows(min_row=2, max_row=max_filas, max_col=hoja.max_column)):
        color_primera_celda = obtener_color_hex(row[0])
        
        # Omitir encabezados azules
        if es_color_azul(color_primera_celda):
            filas_validas.append(False)
            estados.append("Ignorar")
            colores_detectados.append(color_primera_celda)
            continue
            
        filas_validas.append(True)
        
        # Columna donde se suele marcar el color de cancelación (ajustado a la celda 5 o 6)
        cell_f = row[5] if len(row) > 5 else (row[4] if len(row) > 4 else row[0])
        color_f = obtener_color_hex(cell_f)
        colores_detectados.append(color_f)
        
        f_es_cafe = es_color_cafe(color_f)
        total_cafe_fila = sum(1 for cell in row if es_color_cafe(obtener_color_hex(cell)))
        
        if f_es_cafe:
            if total_cafe_fila > 3: 
                estados.append("Cancelada (Por el paciente)")
            else:
                estados.append("Cancelada (Médico no alcanzó)")
        else:
            # AHORA POR DEFECTO ES CONSULTA EFECTIVA
            estados.append("Consulta Efectiva")
            
    # Ajustar longitudes por si pandas leyó distinto a openpyxl
    if len(estados) > len(df): estados = estados[:len(df)]
    elif len(estados) < len(df): estados.extend(["Consulta Efectiva"] * (len(df) - len(estados)))
    
    if len(filas_validas) > len(df): filas_validas = filas_validas[:len(df)]
    elif len(filas_validas) < len(df): filas_validas.extend([True] * (len(df) - len(filas_validas)))
    
    if len(colores_detectados) > len(df): colores_detectados = colores_detectados[:len(df)]
    elif len(colores_detectados) < len(df): colores_detectados.extend(["SIN_COLOR"] * (len(df) - len(colores_detectados)))
            
    df['ESTADO_VISITA'] = estados
    df['ES_VALIDA'] = filas_validas
    df['COLOR_DETECTADO'] = colores_detectados # Guardamos el color para que lo puedas ver
    
    df = df[df['ES_VALIDA'] == True].copy()
    
    columnas_texto = df.select_dtypes(include=['object']).columns
    for col in columnas_texto:
        df[col] = df[col].astype(str).str.strip().str.upper()
        df[col] = df[col].replace({'NAN': None, 'NONE': None, '': None, 'NAT': None})
        
    col_medico = 'MEDICO' if 'MEDICO' in df.columns else df.columns[0]
    
    if col_medico in df.columns:
        df.dropna(subset=[col_medico], inplace=True)
        df = df[~df[col_medico].str.contains(r'AM|PM|RESERVA|MEDICO|PROGRAMAR|ALMUERZO', na=False, regex=True)]
        df[col_medico] = df[col_medico].replace(CORRECCIONES_MEDICOS)
        
    df.dropna(thresh=3, inplace=True)
    df['PRODUCTIVIDAD'] = df['ESTADO_VISITA'].apply(lambda x: 1 if x == "Consulta Efectiva" else 0)
        
    return df

# 3. Interfaz del Tablero
st.image("https://www.virreysolisips.com/wp-content/uploads/2021/08/logo-virrey-solis.png", width=250)
st.title("📊 Tablero de Control y Productividad")
st.markdown("**Gestión de Atención Domiciliaria**")

archivo_subido = st.file_uploader("Carga la programación diaria en Excel (.xlsx)", type=["xlsx"])

if archivo_subido is not None:
    try:
        df = cargar_y_procesar_datos(archivo_subido)
        
        if df.empty:
            st.error("⚠️ El archivo se procesó, pero se quedó sin datos. Revisa la estructura de tu Excel.")
        else:
            col_medico = 'MEDICO' if 'MEDICO' in df.columns else df.columns[0]
            
            # --- PANEL DE DEPURACIÓN (ÁBRELO SI LAS CANCELADAS ESTÁN EN CERO) ---
            with st.expander("🛠️ Ver datos crudos y colores (Depuración)"):
                st.write("Si las cancelaciones no se marcan, busca una cita cancelada aquí y dime qué código aparece en **COLOR_DETECTADO**:")
                st.dataframe(df