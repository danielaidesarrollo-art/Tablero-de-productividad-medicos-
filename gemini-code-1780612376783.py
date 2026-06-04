import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px

# 1. Configuración de la página
st.set_page_config(page_title="Tablero Virrey Solís - Productividad", layout="wide")

# --- COLORES INSTITUCIONALES ---
COLOR_AZUL = "#004B87"
COLOR_VERDE = "#00A650"
COLOR_ALERTA = "#E74C3C" 

CORRECCIONES_MEDICOS = {
    "KARON": "CAROL",
    "KAROL": "CAROL",
    "JHON ERIC": "JHON ERIK",
}

# --- FUNCIONES DE COLOR ---
def obtener_color_hex(cell):
    if cell is None or not hasattr(cell, 'fill') or cell.fill is None:
        return "SIN_COLOR"
    if getattr(cell.fill, 'patternType', None) is None:
        return "SIN_COLOR"
    if getattr(cell.fill, 'start_color', None) is None:
        return "SIN_COLOR"
    
    color_val = str(cell.fill.start_color.index).upper()
    if color_val in ['00000000', '0', '64', 'FFFFFFFF', '00FFFFFF', 'SYSTEM FOREGROUND']:
        return "SIN_COLOR"
        
    return color_val

def es_color_cafe(color_str):
    if color_str == "SIN_COLOR": 
        return False
    rojos_cafes = ['FFFF0000', 'FFC00000', 'FFFFC000', 'FFE26B0A', 'FF974706', 'FFC65911', 'FFED7D31', 'FF806000']
    return any(c in color_str for c in rojos_cafes)

def es_color_azul(color_str):
    if color_str == "SIN_COLOR":
        return False
    azules = ['FF00B0F0', 'FF0070C0', 'FF4472C4', 'FF5B9BD5', 'FF1F497D', 'FFDEEBF7', 'FFBDD7EE', 'FF9BC2E6']
    return any(azul in color_str for azul in azules)

# 2. Procesamiento de Lectura
@st.cache_data
def cargar_y_procesar_datos(archivo_subido):
    df = pd.read_excel(archivo_subido)
    df.columns = df.columns.astype(str).str.strip().str.upper()
    
    archivo_subido.seek(0)
    wb = openpyxl.load_workbook(archivo_subido, data_only=True)
    hoja = wb.active
    
    estados = []
    filas_validas = []
    max_filas = len(df) + 1
    
    for idx, row in enumerate(hoja.iter_rows(min_row=2, max_row=max_filas, max_col=hoja.max_column)):
        color_primera_celda = obtener_color_hex(row[0])
        
        if es_color_azul(color_primera_celda):
            filas_validas.append(False)
            estados.append("IGNORAR")
            continue
            
        filas_validas.append(True)
        
        # Buscar el color en la columna F (índice 5) o cercanas
        cell_f = row[5] if len(row) > 5 else (row[4] if len(row) > 4 else row[0])
        color_f = obtener_color_hex(cell_f)
        
        f_es_cafe = es_color_cafe(color_f)
        total_cafe_fila = sum(1 for cell in row if es_color_cafe(obtener_color_hex(cell)))
        
        if f_es_cafe:
            if total_cafe_fila > 3: 
                estados.append("CANCELADA (POR EL PACIENTE)")
            else:
                estados.append("CANCELADA (MÉDICO NO ALCANZÓ)")
        else:
            estados.append("CONSULTA EFECTIVA")
            
    if len(estados) > len(df): estados = estados[:len(df)]
    elif len(estados) < len(df): estados.extend(["CONSULTA EFECTIVA"] * (len(df) - len(estados)))
    
    if len(filas_validas) > len(df): filas_validas = filas_validas[:len(df)]
    elif len(filas_validas) < len(df): filas_validas.extend([True] * (len(df) - len(filas_validas)))
            
    df['ESTADO_VISITA'] = estados
    df['ES_VALIDA'] = filas_validas
    
    df = df[df['ES_VALIDA'] == True].copy()
    
    # Limpieza de textos y Nulos
    columnas_texto = df.select_dtypes(include=['object']).columns
    for col in columnas_texto:
        df[col] = df[col].astype(str).str.strip().str.upper()
        df[col] = df[col].replace({'NAN': None, 'NONE': None, '': None, 'NAT': None})
        
    return df

# 3. Interfaz del Tablero
st.image("https://www.virreysolisips.com/wp-content/uploads/2021/08/logo-virrey-solis.png", width=250)
st.title("📊 Tablero de Control y Productividad")
st.markdown("**Gestión de Atención Domiciliaria**")

archivo_subido = st.file_uploader("Carga la programación diaria en Excel (.xlsx)", type=["xlsx"])

if archivo_subido is not None:
    try:
        df = cargar_y_procesar_datos(archivo_subido)
        
        if df.empty:
            st.error("⚠️ El archivo se procesó, pero se quedó sin datos. Revisa la estructura de tu Excel.")
        else:
            # --- LIMPIEZA DE MÉDICOS ---
            col_medico = 'MEDICO' if 'MEDICO' in df.columns else df.columns[0]
            if col_medico in df.columns:
                df = df[df[col_medico].notna()]
                df = df[~df[col_medico].str.contains(r'AM|PM|RESERVA|MEDICO|PROGRAMAR|ALMUERZO', na=False, regex=True)]
                df[col_medico] = df[col_medico].replace(CORRECCIONES_MEDICOS)
            
            # ================= BARRA LATERAL (SIDEBAR) =================
            st.sidebar.header("1. Configuración de Datos")
            
            # Filtro de Paciente (Para borrar vacíos)
            sugerencia_idx = 0
            for i, col in enumerate(df.columns):
                if any(palabra in col for palabra in ['PACIENTE', 'NOMBRE', 'DOC', 'CEDULA']):
                    sugerencia_idx = i
                    break
            col_paciente = st.sidebar.selectbox("Columna de Paciente (Borra vacíos)", options=df.columns, index=sugerencia_idx)
            
            if col_paciente:
                df = df[df[col_paciente].notna()]
            
            # --- NUEVO: FILTRO POR COLUMNA D (MES) ---
            st.sidebar.header("2. Filtros de Tiempo")
            
            # Verificamos que el Excel tenga al menos 4 columnas (A, B, C, D)
            if len(df.columns) > 3:
                # La columna D es el índice 3
                col_mes = df.columns[3] 
                
                # Creamos una lista de los meses únicos ignorando vacíos
                meses_disponibles = sorted([m for m in df[col_mes].unique() if str(m) not in ['NAN', 'NONE', '']])
                
                if meses_disponibles:
                    # Por defecto mostramos todo lo disponible
                    mes_seleccionado = st.sidebar.multiselect(
                        f"Filtrar por Mes (Columna: {col_mes})", 
                        options=meses_disponibles, 
                        default=meses_disponibles
                    )
                    
                    if mes_seleccionado:
                        df = df[df[col_mes].isin(mes_seleccionado)]
                else:
                    st.sidebar.info("La Columna D está vacía.")
            else:
                st.sidebar.warning("El archivo no tiene suficientes columnas (Columna D).")

            # Filtro de Médico
            st.sidebar.header("3. Filtros de Personal")
            medicos_sel = st.sidebar.multiselect("Filtrar por Médico", options=df[col_medico].dropna().unique())
            if medicos_sel: 
                df = df[df[col_medico].isin(medicos_sel)]

            # ================= CÁLCULO FINAL Y GRÁFICAS =================
            df['PRODUCTIVIDAD'] = df['ESTADO_VISITA'].apply(lambda x: 1 if "EFECTIVA" in str(x) else 0)
                
            st.markdown("### 📈 Indicadores Globales")
            c