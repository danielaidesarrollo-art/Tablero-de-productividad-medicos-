import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px

# 1. Configuración de la página del tablero (Canvas)
st.set_page_config(page_title="Tablero de Productividad Médica", layout="wide")

# Función auxiliar para identificar si una celda es de color café (excluyendo amarillos y vacíos)
def es_color_cafe(cell):
    if not hasattr(cell, 'fill') or not cell.fill or not cell.fill.start_color:
        return False
    color_str = str(cell.fill.start_color.index).upper()
    
    # Filtrar celdas sin color (blancas o transparentes)
    if color_str in ['00000000', '000000', 'NONE', '']:
        return False
    
    # Filtrar tonos amarillos comunes en Excel (ej: FFFF00, FFF2CC, etc.)
    if 'FF' in color_str and any(y in color_str for y in ['FFFF00', 'FFF2CC', 'FFFF99', 'FFFFCC', 'FFEB9C', 'FFFFC000']):
        return False
        
    return True

# 2. Función para cargar y procesar los datos con la nueva lógica
@st.cache_data
def cargar_y_procesar_datos(archivo_subido):
    df = pd.read_excel(archivo_subido)
    df.columns = df.columns.str.strip() # Limpiar espacios en los nombres
    
    # Rebobinar el archivo para openpyxl
    archivo_subido.seek(0)
    wb = openpyxl.load_workbook(archivo_subido, data_only=True)
    hoja = wb.active
    
    estados = []
    max_filas = len(df) + 1
    
    # Iterar sobre las filas reales de Excel
    for row in hoja.iter_rows(min_row=2, max_row=max_filas, max_col=hoja.max_column):
        # Columna F es el índice 5 (DOCUMENTO)
        cell_f = row[5] 
        
        f_es_cafe = es_color_cafe(cell_f)
        total_cafe_fila = sum(1 for cell in row if es_color_cafe(cell))
        
        # Nueva lógica de negocio definida por el usuario
        if f_es_cafe:
            # Si toda o la mayor parte de la fila está en café, canceló el paciente
            if total_cafe_fila > 3: 
                estados.append("Cancelada (Por el paciente)")
            else:
                # Si solo está café en la columna F, el médico no alcanzó
                estados.append("Cancelada (Médico no alcanzó)")
        else:
            # Sin color o amarillo en columna F = Consulta Efectiva
            estados.append("Consulta Efectiva")
            
    # Ajuste preventivo de tamaños de listas
    if len(estados) > len(df):
        estados = estados[:len(df)]
    elif len(estados) < len(df):
        estados.extend(["Consulta Efectiva"] * (len(df) - len(estados)))
            
    df['Estado_Visita'] = estados
    
    # Sumar a la productividad únicamente las Consultas Efectivas
    df['Productividad'] = df['Estado_Visita'].apply(
        lambda x: 1 if x == "Consulta Efectiva" else 0
    )
    
    return df

# 3. Interfaz de usuario del Tablero
st.title("📊 Tablero de Productividad Médica y Control de Metas")
st.markdown("Carga el archivo de programación en formato **Excel (.xlsx)** para actualizar las métricas.")

archivo_subido = st.file_uploader("Selecciona tu archivo Excel", type=["xlsx"])

if archivo_subido is not None:
    try:
        df_procesado = cargar_y_procesar_datos(archivo_subido)
        
        # Identificar dinámicamente columnas clave
        col_medico = 'MEDICO' if 'MEDICO' in df_procesado.columns else df_procesado.columns[0]
        col_genero = next((c for c in df_procesado.columns if c.upper() in ['GENERO', 'GÉNERO', 'SEXO', 'SEX']), None)
        
        # Filtros laterales interactivos
        st.sidebar.header("Filtros del Tablero")
        medicos_sel = st.sidebar.multiselect("Filtrar por Médico", options=df_procesado[col_medico].dropna().unique())
        
        if medicos_sel:
            df_procesado = df_procesado[df_procesado[col_medico].isin(medicos_sel)]
            
        # 4. KPIs Principales
        st.markdown("### 📈 Indicadores de Rendimiento")
        c1, c2, c3 = st.columns(3)
        
        total_visitas = len(df_procesado)
        total_efectivas = df_procesado['Productividad'].sum()
        rendimiento = (total_efectivas / total_visitas) * 100 if total_visitas > 0 else 0
        
        c1.metric("Visitas Asignadas", total_visitas)
        c2.metric("Consultas Efectivas Realizadas", total_efectivas)
        c3.metric("Porcentaje de Efectividad", f"{rendimiento:.1f}%")
        
        st.markdown("---")
        
        # 5. Gráficos de Productividad y Metas
        colA, colB = st.columns(2)
        
        with colA:
            st.subheader("Productividad Total vs Meta (162 Consultas)")
            df_medicos = df_procesado.groupby(col_medico)['Productividad'].sum().reset_index()
            df_medicos = df_medicos.sort_values(by='Productividad', ascending=False)
            
            fig_bar = px.bar(
                df_medicos, x=col_medico, y='Productividad', text='Productividad',
                color='Productividad', color_continuous_scale='Greens'
            )
            # Añadir línea de meta horizontal en 162 consultas
            fig_bar.add_hline(y=162, line_dash="dash", line_color="red", annotation_text="Meta: 162", annotation_position="top left")
            st.plotly_chart(fig_bar, use_container_width=True)
            
        with colB:
            st.subheader("Evolución de Consultas en el Tiempo")
            # Asegurar formato de fecha para la gráfica de líneas
            df_procesado['FECHA_FORMATO'] = pd.to_datetime(df_procesado['FECHA'], errors='coerce').dt.strftime('%Y-%m-%d')
            df_linea = df_procesado.groupby(['FECHA_FORMATO', col_medico])['Productividad'].sum().reset_index()
            df_linea = df_linea.sort_values('FECHA_FORMATO')
            
            fig_linea = px.line(
                df_linea, x='FECHA_FORMATO', y='Productividad', color=col_medico,
                title="Comparativa Diaria entre Médicos", markers=True
            )
            st.plotly_chart(fig_linea, use_container_width=True)
            
        # 6. Gráficos Secundarios: Estados y Género
        st.markdown("---")
        colC, colD = st.columns(2)
        
        with colC:
            st.subheader("Distribución de Motivos de Visita")
            fig_pie = px.pie(df_procesado, names='Estado_Visita', hole=0.4, color_discrete_sequence=['#2ecc71', '#e67e22', '#e74c3c'])
            st.plotly_chart(fig_pie, use_container_width=True)
            
        with colD:
            st.subheader("Análisis por Género de los Pacientes")
            if col_genero:
                fig_gen = px.pie(df_procesado, names=col_genero, hole=0.4, color_discrete_sequence=['#3498db', '#e74c3c', '#9b59b6'])
                st.plotly_chart(fig_gen, use_container_width=True)
            else:
                st.info("Para ver el gráfico de género, asegúrate de tener una columna llamada 'GENERO' o 'SEXO' en tu archivo Excel.")
                
        # 7. Vista de datos final
        st.markdown("### 📋 Datos Analizados")
        st.dataframe(df_procesado, use_container_width=True)

    except Exception as e:
        st.error(f"Error en la estructura del archivo. Verifica que corresponda al formato correcto. Detalle: {e}")