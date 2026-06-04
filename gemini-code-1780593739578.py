import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px

# 1. Configuración de la página del tablero
st.set_page_config(page_title="Tablero de Productividad Médica", layout="wide")

# Función auxiliar para identificar si una celda es de color café
def es_color_cafe(cell):
    if not hasattr(cell, 'fill') or not cell.fill or not cell.fill.start_color:
        return False
    color_str = str(cell.fill.start_color.index).upper()
    
    # Filtrar celdas sin color (blancas o transparentes)
    if color_str in ['00000000', '000000', 'NONE', '']:
        return False
    
    # Filtrar tonos amarillos comunes en Excel
    if 'FF' in color_str and any(y in color_str for y in ['FFFF00', 'FFF2CC', 'FFFF99', 'FFFFCC', 'FFEB9C', 'FFFFC000']):
        return False
        
    return True

# 2. Función para cargar, procesar y LIMPIAR los datos
@st.cache_data
def cargar_y_procesar_datos(archivo_subido):
    # A. Lectura inicial
    df = pd.read_excel(archivo_subido)
    
    # --- LIMPIEZA 1: Estandarizar nombres de columnas ---
    df.columns = df.columns.astype(str).str.strip().str.upper()
    
    # B. Extraer colores con openpyxl
    archivo_subido.seek(0)
    wb = openpyxl.load_workbook(archivo_subido, data_only=True)
    hoja = wb.active
    
    estados = []
    max_filas = len(df) + 1
    
    # Iterar sobre las filas
    for row in hoja.iter_rows(min_row=2, max_row=max_filas, max_col=hoja.max_column):
        # Asegurarnos de que existe la columna F (índice 5)
        cell_f = row[5] if len(row) > 5 else None 
        
        f_es_cafe = es_color_cafe(cell_f)
        total_cafe_fila = sum(1 for cell in row if es_color_cafe(cell))
        
        # Lógica de negocio
        if f_es_cafe:
            if total_cafe_fila > 3: 
                estados.append("Cancelada (Por el paciente)")
            else:
                estados.append("Cancelada (Médico no alcanzó)")
        else:
            estados.append("Consulta Efectiva")
            
    # Ajustar desfases de longitud
    if len(estados) > len(df):
        estados = estados[:len(df)]
    elif len(estados) < len(df):
        estados.extend(["Consulta Efectiva"] * (len(df) - len(estados)))
            
    # Asignar estados y productividad al DataFrame antes de limpiar filas
    df['ESTADO_VISITA'] = estados
    df['PRODUCTIVIDAD'] = df['ESTADO_VISITA'].apply(
        lambda x: 1 if x == "Consulta Efectiva" else 0
    )
    
    # --- LIMPIEZA 2: Eliminar filas basura/fantasma ---
    # Busca la columna del médico y documento para saber si la fila es válida
    col_medico_val = 'MEDICO' if 'MEDICO' in df.columns else df.columns[0]
    col_doc_val = 'DOCUMENTO' if 'DOCUMENTO' in df.columns else None
    
    columnas_validacion = [col_medico_val]
    if col_doc_val: columnas_validacion.append(col_doc_val)
    
    # Borrar la fila entera si NO tiene Médico ni Documento (eran filas en blanco)
    df.dropna(subset=columnas_validacion, how='all', inplace=True)
    
    # --- LIMPIEZA 3: Estandarizar textos (Mayúsculas y quitar espacios extra) ---
    columnas_texto = df.select_dtypes(include=['object']).columns
    for col in columnas_texto:
        df[col] = df[col].astype(str).str.strip().str.upper()
        # Convertir los textos vacíos y "NAN" que generó pandas de vuelta a nulos reales
        df[col] = df[col].replace({'NAN': None, 'NONE