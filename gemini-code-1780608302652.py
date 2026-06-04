import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px

# 1. Configuración de la página
st.set_page_config(page_title="Tablero Virrey Solís - Productividad", layout="wide")

# --- COLORES INSTITUCIONALES VIRREY SOLIS ---
COLOR_AZUL = "#004B87"
COLOR_VERDE = "#00A650"
COLOR_ALERTA = "#E74C3C" 
PALETA_VS = [COLOR_AZUL, COLOR_VERDE, "#85C1E9", "#A9DFBF", "#154360"]

# --- DICCIONARIO DE CORRECCIÓN DE NOMBRES ---
CORRECCIONES_MEDICOS = {
    "KARON": "CAROL",
    "KAROL": "CAROL",
    "JHON ERIC": "JHON ERIK",
}

# Funciones para colores
def obtener_color_hex(cell):
    if not hasattr(cell, 'fill') or not cell.fill or not cell.fill.start_color:
        return ""
    return str(cell.fill.start_color.index).upper()

def es_color_cafe(color_str):
    if color_str in ['00000000', '000000', 'NONE', '']: return False
    if 'FF' in color_str and any(y in color_str for y in ['FFFF00', 'FFF2CC', 'FFFF99', 'FFFFCC', 'FFEB9C', 'FFFFC000']):
        return False
    if es_color_azul(color_str): return False
    return True

def es_color_azul(color_str):
    azules = ['FF00B0F0', 'FF0070C0', 'FF4472C4', 'FF5B9BD5', 'FF1F497D', 'FFDEEBF7', 'FFBDD7EE', 'FF9BC2E6']
    return any(azul in color_str for azul in azules)

# 2. Procesamiento y Limpieza
@st.cache_data
def cargar_y_procesar_datos(archivo_subido):
    df = pd.read_excel(archivo_subido)
    df.columns = df.columns.astype(str).str.strip().str.upper()
    
    archivo_subido.seek(0)
    wb = openpyxl.load_workbook(archivo_subido, data_only=True)
    hoja = wb.active
    
    estados = []
    filas_validas = []
    max_filas = len(df) + 1
    
    for idx, row in enumerate(hoja.iter_rows(min_row=2, max_row=max_filas, max_col=hoja.max_column)):
        color_primera_celda = obtener_color_hex(row[0])
        if es_color_azul(color_primera_celda):
            filas_validas.append(False)
            estados.append("Ignorar")
            continue
            
        filas_validas.append(True)
        cell_f = row[5] if len(row) > 5 else None 
        color_f = obtener_color_hex(cell_f)
        
        f_es_cafe = es_color_cafe(color_f)
        total_cafe_fila = sum(1 for cell in row if es_color_cafe(obtener_color_hex(cell)))
        
        if f_es_cafe:
            if total_cafe_fila > 3: 
                estados.append("Cancelada (Por el paciente)")
            else:
                estados.append("Cancelada (Médico no alcanzó)")
        else:
            estados.append("Consulta Efectiva")
            
    if len(estados) > len(df): estados = estados[:len(df)]
    elif len(estados) < len(df): estados.extend(["Consulta Efectiva"] * (len(df) - len(estados)))
    
    if len(filas_validas) > len(df): filas_validas = filas_validas[:len(df)]
    elif len(filas_validas) < len(df): filas_validas.extend([True] * (len(df) - len(filas_validas)))
            
    df['ESTADO_VISITA'] = estados
    df['ES_VALIDA'] = filas_validas
    df = df[df['ES_VALIDA'] == True].copy()
    
    # Limpieza de textos
    columnas_texto = df.select_dtypes(include=['object']).columns
    for col in columnas_texto:
        df[col] = df[col].astype(str).str.strip().str.upper()
        df[col] = df[col].replace({'NAN': None, 'NONE': None, '': None, 'NAT': None})
        
    # --- FILTROS DE CORRECCIÓN ---
    col_medico = 'MEDICO' if 'MEDICO' in df.columns else df.columns[0]
    
    if col_medico in df.columns:
        # Eliminar filas donde el médico sea nulo
        df.dropna(subset=[col_medico], inplace=True)
        # Eliminar falsos médicos (notas, horas, reservas)
        df = df[~df[col_medico].str.contains(r'AM|PM|RESERVA|MEDICO|PROGRAMAR|ALMUERZO', na=False, regex=True)]
        # Aplicar correcciones ortográficas
        df[col_medico] = df[col_medico].replace(CORRECCIONES_MEDICOS)
        
    # Eliminar turnos en blanco (donde no hay paciente)
    col_paciente = next((c for c in df.columns if 'PACIENTE' in c or 'NOMBRE' in c or 'DOC' in c or 'CEDULA' in c), None)
    if col_paciente:
        df.dropna(subset=[col_paciente], inplace=True)
        
    # Calcular productividad solo sobre lo que quedó
    df['PRODUCTIVIDAD'] = df['ESTADO_VISITA'].apply(lambda x: 1 if x == "Consulta Efectiva" else 0)
        
    return df

# 3. Interfaz del Tablero
st.image("https://www.virreysolisips.com/wp-content/uploads/2021/08/logo-virrey-solis.png", width=250)
st.title("📊 Tablero de Control y Productividad")
st.markdown("**Gestión de Atención Domiciliaria**")

archivo_subido = st.file_uploader("Carga la programación diaria en Excel (.xlsx)", type=["xlsx"])

if archivo_subido is not None:
    try:
        df = cargar_y_procesar_datos(archivo_subido)
        col_medico = 'MEDICO' if 'MEDICO' in df.columns else df.columns[0]
        
        st.sidebar.header("Filtros de Análisis")
        medicos_sel = st.sidebar.multiselect("Filtrar por Médico", options=df[col_medico].dropna().unique())
        if medicos_sel: df = df[df[col_medico].isin(medicos_sel)]
            
        st.markdown("### 📈 Indicadores Globales")
        c1, c2, c3 = st.columns(3)
        total_visitas = len(df)
        total_efectivas = df['PRODUCTIVIDAD'].sum()
        rendimiento = (total_efectivas / total_visitas) * 100 if total_visitas > 0 else 0
        
        c1.metric("Visitas Asignadas", total_visitas)
        c2.metric("Consultas Efectivas", total_efectivas)
        c3.metric("Efectividad", f"{rendimiento:.1f}%")
        st.markdown("---")
        
        st.subheader("Productividad por Médico vs Meta (162)")
        df_med = df.groupby(col_medico)['PRODUCTIVIDAD'].sum().reset_index()
        fig_bar = px.bar(
            df_med.sort_values('PRODUCTIVIDAD', ascending=False), 
            x=col_medico, y='PRODUCTIVIDAD', text='PRODUCTIVIDAD',
            color_discrete_sequence=[COLOR_VERDE]
        )
        fig_bar.add_hline(y=162, line_dash="dash", line_color=COLOR_AZUL, annotation_text="Meta 162")
        st.plotly_chart(fig_bar, use_container_width=True)
        
        st.markdown("---")
        st.markdown(f"<h3 style='color: {COLOR_AZUL};'>🚨 Análisis de Cancelaciones</h3>", unsafe_allow_html=True)
        
        df_canceladas = df[df['ESTADO_VISITA'].str.contains("Cancelada", na=False)]
        
        if not df_canceladas.empty:
            colA, colB = st.columns(2)
            
            with colA:
                st.subheader("⏱️ Jornada con más cancelaciones")
                col_jornada = next((c for c in df.columns if 'JORNADA' in c), None)
                
                if col_jornada:
                    df_jornada = df_canceladas.groupby(col_jornada).size().reset_index(name='Cantidad')
                    fig_jornada = px.bar(
                        df_jornada.sort_values('Cantidad', ascending=False), 
                        x=col_jornada, y='Cantidad', text='Cantidad',
                        color_discrete_sequence=[COLOR_ALERTA]
                    )
                    fig_jornada.update_layout(xaxis_title="Jornada", yaxis_title="Total Cancelaciones")
                    st.plotly_chart(fig_jornada, use_container_width=True)
                else: