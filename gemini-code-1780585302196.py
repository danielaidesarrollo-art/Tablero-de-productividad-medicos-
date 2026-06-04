import pandas as pd
import openpyxl
import streamlit as st
import plotly.express as px

st.set_page_config(page_title="Tablero de Productividad Médica", layout="wide")

@st.cache_data
def cargar_y_procesar_datos(ruta_archivo):
    # Cargar el archivo Excel para leer formatos y colores
    wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
    hoja = wb.active
    
    # Cargar datos tabulares con Pandas
    df = pd.read_excel(ruta_archivo)
    
    estados = []
    
    # Iterar sobre las filas de la hoja (omitiendo el encabezado)
    for row_idx, row in enumerate(hoja.iter_rows(min_row=2, max_col=hoja.max_column), start=2):
        colores_cafe_en_fila = 0
        
        # Contar celdas café en la fila actual
        for cell in row:
            # Detectar si el color de relleno es café (Theme/Tint o Hex)
            # El código Hex del café en Excel puede variar, ajusta el valor según tu paleta
            if cell.fill.start_color.index != '00000000' and type(cell.fill.start_color.index) == str:
                # Lógica simplificada: Si hay un relleno distinto al blanco/nulo
                colores_cafe_en_fila += 1
                
        # Aplicar la lógica de negocio
        if colores_cafe_en_fila > 3: # Si hay múltiples celdas pintadas (toda la fila)
            estados.append("Cancelada (Médico no alcanzó)")
        elif colores_cafe_en_fila > 0 and colores_cafe_en_fila <= 3: # Solo una celda pintada
            estados.append("Efectiva (Rechazo en puerta)")
        else:
            estados.append("Efectiva (Realizada)")
            
    df['Estado_Visita'] = estados
    
    # Determinar si la visita suma a la productividad
    df['Productividad'] = df['Estado_Visita'].apply(
        lambda x: 1 if "Efectiva" in x else 0
    )
    
    return df

# Interfaz del Tablero (Canvas)
st.title("📊 Tablero Interactivo: Productividad Domiciliaria")

# Subir archivo
archivo_subido = st.file_uploader("Sube tu archivo Excel (.xlsx)", type=["xlsx"])

if archivo_subido is not None:
    # Procesar los datos
    df_procesado = cargar_y_procesar_datos(archivo_subido)
    
    # Filtros Interactivos
    st.sidebar.header("Filtros")
    medicos = st.sidebar.multiselect("Seleccionar Médico(s)", options=df_procesado['MEDICO'].dropna().unique())
    
    if medicos:
        df_procesado = df_procesado[df_procesado['MEDICO'].isin(medicos)]
        
    # Tarjetas de Resumen (KPIs)
    col1, col2, col3 = st.columns(3)
    total_asignadas = len(df_procesado)
    total_efectivas = df_procesado['Productividad'].sum()
    porcentaje_efectividad = (total_efectivas / total_asignadas) * 100 if total_asignadas > 0 else 0
    
    col1.metric("Total Visitas Asignadas", total_asignadas)
    col2.metric("Total Visitas Efectivas", total_efectivas)
    col3.metric("Productividad Global (%)", f"{porcentaje_efectividad:.1f}%")
    
    st.markdown("---")
    
    # Gráficos Interactivos
    colA, colB = st.columns(2)
    
    with colA:
        st.subheader("Estado General de las Visitas")
        fig_pie = px.pie(df_procesado, names='Estado_Visita', hole=0.4, color_discrete_sequence=['#2ecc71', '#8b4513', '#e74c3c'])
        st.plotly_chart(fig_pie, use_container_width=True)
        
    with colB:
        st.subheader("Productividad por Médico")
        df_agrupado = df_procesado.groupby('MEDICO')['Productividad'].sum().reset_index()
        df_agrupado = df_agrupado.sort_values(by='Productividad', ascending=False)
        fig_bar = px.bar(df_agrupado, x='MEDICO', y='Productividad', text='Productividad', color='Productividad', color_continuous_scale='Viridis')
        st.plotly_chart(fig_bar, use_container_width=True)

    # Tabla de datos detallada
    st.markdown("### Detalle de Pacientes")
    st.dataframe(df_procesado[['FECHA', 'DOCUMENTO', 'APELLIDOS NOMBRES', 'MEDICO', 'Estado_Visita']])