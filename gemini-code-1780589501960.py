import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px

# 1. Configuración de la página
st.set_page_config(page_title="Tablero de Productividad Médica", layout="wide")

@st.cache_data
def cargar_y_procesar_datos(archivo_subido):
    # Leer el archivo con Pandas
    df = pd.read_excel(archivo_subido)
    df.columns = df.columns.str.strip() # Limpiar espacios en los nombres
    
    # "Rebobinar" el archivo al inicio
    archivo_subido.seek(0)
    
    # Leer el archivo con openpyxl para detectar los colores
    wb = openpyxl.load_workbook(archivo_subido, data_only=True)
    hoja = wb.active
    
    estados = []
    
    # EL ARREGLO: Límite exacto de filas (Datos reales + 1 del encabezado)
    max_filas = len(df) + 1
    
    # Iterar solo hasta donde hay datos reales
    for row in hoja.iter_rows(min_row=2, max_row=max_filas, max_col=hoja.max_column):
        colores_cafe_en_fila = 0
        
        # Contar celdas coloreadas
        for cell in row:
            if hasattr(cell, 'fill') and cell.fill.start_color.index != '00000000' and type(cell.fill.start_color.index) == str:
                colores_cafe_en_fila += 1
                
        # Aplicar la lógica de colores
        if colores_cafe_en_fila > 3: 
            estados.append("Cancelada (Médico no alcanzó)")
        elif colores_cafe_en_fila > 0 and colores_cafe_en_fila <= 3: 
            estados.append("Efectiva (Rechazo en puerta)")
        else: 
            estados.append("Efectiva (Realizada)")
            
    # Seguro adicional: igualar tamaños por si hay algún desfase
    if len(estados) > len(df):
        estados = estados[:len(df)]
    elif len(estados) < len(df):
        estados.extend(["Efectiva (Realizada)"] * (len(df) - len(estados)))
            
    df['Estado_Visita'] = estados
    
    # Determinar si la visita suma a la productividad
    df['Productividad'] = df['Estado_Visita'].apply(
        lambda x: 1 if "Efectiva" in x else 0
    )
    
    return df
    
    # Leer el archivo con openpyxl para detectar los colores
    wb = openpyxl.load_workbook(archivo_subido, data_only=True)
    hoja = wb.active
    
    estados = []
    
    # Iterar sobre las filas (omitiendo la fila 1 de encabezados)
    for row_idx, row in enumerate(hoja.iter_rows(min_row=2, max_col=hoja.max_column), start=2):
        colores_cafe_en_fila = 0
        
        # Contar celdas coloreadas en la fila actual
        for cell in row:
            # Detectar si el color de relleno es distinto de nulo (blanco/transparente)
            if hasattr(cell, 'fill') and cell.fill.start_color.index != '00000000' and type(cell.fill.start_color.index) == str:
                colores_cafe_en_fila += 1
                
        # Aplicar la lógica de negocio según los colores
        if colores_cafe_en_fila > 3: # Toda la fila pintada
            estados.append("Cancelada (Médico no alcanzó)")
        elif colores_cafe_en_fila > 0 and colores_cafe_en_fila <= 3: # Solo algunas celdas pintadas
            estados.append("Efectiva (Rechazo en puerta)")
        else: # Ninguna celda pintada
            estados.append("Efectiva (Realizada)")
            
    df['Estado_Visita'] = estados
    
    # Determinar si la visita suma a la productividad (1 si es efectiva, 0 si fue cancelada)
    df['Productividad'] = df['Estado_Visita'].apply(
        lambda x: 1 if "Efectiva" in x else 0
    )
    
    return df

# 3. Interfaz del Tablero (UI)
st.title("📊 Tablero Interactivo: Productividad Domiciliaria")
st.markdown("Sube la programación en formato **Excel (.xlsx)** para visualizar la productividad de los médicos.")

# Subir archivo
archivo_subido = st.file_uploader("Selecciona tu archivo Excel", type=["xlsx"])

if archivo_subido is not None:
    try:
        # Procesar los datos
        df_procesado = cargar_y_procesar_datos(archivo_subido)
        
        # Identificar la columna del médico (asume 'MEDICO' u otra similar, ajusta si es necesario)
        columna_medico = 'MEDICO' if 'MEDICO' in df_procesado.columns else df_procesado.columns[0]
        
        # Filtros en la barra lateral
        st.sidebar.header("Filtros")
        medicos_seleccionados = st.sidebar.multiselect(
            "Seleccionar Médico(s)", 
            options=df_procesado[columna_medico].dropna().unique()
        )
        
        # Aplicar filtro si se seleccionó algún médico
        if medicos_seleccionados:
            df_procesado = df_procesado[df_procesado[columna_medico].isin(medicos_seleccionados)]
            
        # 4. Tarjetas de Resumen (KPIs)
        st.markdown("### Resumen Global")
        col1, col2, col3 = st.columns(3)
        
        total_asignadas = len(df_procesado)
        total_efectivas = df_procesado['Productividad'].sum()
        porcentaje_efectividad = (total_efectivas / total_asignadas) * 100 if total_asignadas > 0 else 0
        
        col1.metric("Total Visitas Asignadas", total_asignadas)
        col2.metric("Total Visitas Efectivas", total_efectivas)
        col3.metric("Productividad Global (%)", f"{porcentaje_efectividad:.1f}%")
        
        st.markdown("---")
        
        # 5. Gráficos Interactivos
        colA, colB = st.columns(2)
        
        with colA:
            st.subheader("Estado General de las Visitas")
            fig_pie = px.pie(
                df_procesado, 
                names='Estado_Visita', 
                hole=0.4, 
                color_discrete_sequence=['#2ecc71', '#e74c3c', '#e67e22']
            )
            st.plotly_chart(fig_pie, use_container_width=True)
            
        with colB:
            st.subheader("Productividad Efectiva por Médico")
            df_agrupado = df_procesado.groupby(columna_medico)['Productividad'].sum().reset_index()
            df_agrupado = df_agrupado.sort_values(by='Productividad', ascending=False)
            
            fig_bar = px.bar(
                df_agrupado, 
                x=columna_medico, 
                y='Productividad', 
                text='Productividad', 
                color='Productividad', 
                color_continuous_scale='Blues'
            )
            st.plotly_chart(fig_bar, use_container_width=True)

        # 6. Tabla de datos detallada
        st.markdown("### 📋 Detalle de Pacientes Procesados")
        # Mostramos las columnas más relevantes, si existen
        columnas_mostrar = [c for c in ['FECHA', 'DOCUMENTO', 'APELLIDOS', 'NOMBRES', columna_medico, 'Estado_Visita'] if c in df_procesado.columns]
        
        if len(columnas_mostrar) > 0:
            st.dataframe(df_procesado[columnas_mostrar], use_container_width=True)
        else:
            st.dataframe(df_procesado, use_container_width=True)

    except Exception as e:
        st.error(f"Ocurrió un error al procesar el archivo. Por favor verifica que sea un Excel válido. Detalle del error: {e}")
