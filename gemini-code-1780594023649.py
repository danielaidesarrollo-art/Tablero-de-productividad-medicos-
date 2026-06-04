import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px

# 1. Configuración de la página del tablero
st.set_page_config(page_title="Tablero de Productividad Médica", layout="wide")

# Función auxiliar para identificar si una celda es de color café
def es_color_cafe(cell):
    if not hasattr(cell, 'fill') or not cell.fill or not cell.fill.start_color:
        return False
    color_str = str(cell.fill.start_color.index).upper()
    
    # Filtrar celdas sin color (blancas o transparentes)
    if color_str in ['00000000', '000000', 'NONE', '']:
        return False
    
    # Filtrar tonos amarillos comunes en Excel
    if 'FF' in color_str and any(y in color_str for y in ['FFFF00', 'FFF2CC', 'FFFF99', 'FFFFCC', 'FFEB9C', 'FFFFC000']):
        return False
        
    return True

# 2. Función para cargar, procesar y LIMPIAR los datos
@st.cache_data
def cargar_y_procesar_datos(archivo_subido):
    # A. Lectura inicial
    df = pd.read_excel(archivo_subido)
    
    # --- LIMPIEZA 1: Estandarizar nombres de columnas ---
    df.columns = df.columns.astype(str).str.strip().str.upper()
    
    # B. Extraer colores con openpyxl
    archivo_subido.seek(0)
    wb = openpyxl.load_workbook(archivo_subido, data_only=True)
    hoja = wb.active
    
    estados = []
    max_filas = len(df) + 1
    
    # Iterar sobre las filas
    for row in hoja.iter_rows(min_row=2, max_row=max_filas, max_col=hoja.max_column):
        # Asegurarnos de que existe la columna F (índice 5)
        cell_f = row[5] if len(row) > 5 else None 
        
        f_es_cafe = es_color_cafe(cell_f)
        total_cafe_fila = sum(1 for cell in row if es_color_cafe(cell))
        
        # Lógica de negocio
        if f_es_cafe:
            if total_cafe_fila > 3: 
                estados.append("Cancelada (Por el paciente)")
            else:
                estados.append("Cancelada (Médico no alcanzó)")
        else:
            estados.append("Consulta Efectiva")
            
    # Ajustar desfases de longitud
    if len(estados) > len(df):
        estados = estados[:len(df)]
    elif len(estados) < len(df):
        estados.extend(["Consulta Efectiva"] * (len(df) - len(estados)))
            
    # Asignar estados y productividad al DataFrame antes de limpiar filas
    df['ESTADO_VISITA'] = estados
    df['PRODUCTIVIDAD'] = df['ESTADO_VISITA'].apply(
        lambda x: 1 if x == "Consulta Efectiva" else 0
    )
    
    # --- LIMPIEZA 2: Eliminar filas basura/fantasma ---
    col_medico_val = 'MEDICO' if 'MEDICO' in df.columns else df.columns[0]
    col_doc_val = 'DOCUMENTO' if 'DOCUMENTO' in df.columns else None
    
    columnas_validacion = [col_medico_val]
    if col_doc_val: columnas_validacion.append(col_doc_val)
    
    # Borrar la fila entera si NO tiene Médico ni Documento
    df.dropna(subset=columnas_validacion, how='all', inplace=True)
    
    # --- LIMPIEZA 3: Estandarizar textos (Mayúsculas y quitar espacios extra) ---
    columnas_texto = df.select_dtypes(include=['object']).columns
    for col in columnas_texto:
        df[col] = df[col].astype(str).str.strip().str.upper()
        # LA LÍNEA QUE CAUSÓ EL ERROR ESTÁ CORREGIDA AQUÍ ABAJO:
        df[col] = df[col].replace({'NAN': None, 'NONE': None, '': None, 'NAT': None})
        
    # --- LIMPIEZA 4: Formatear Fechas ---
    if 'FECHA' in df.columns:
        df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce')
        
    return df

# 3. Interfaz de usuario del Tablero
st.title("📊 Tablero de Productividad Médica")
st.markdown("Carga tu **Excel (.xlsx)**. El sistema limpiará automáticamente los datos, unificará mayúsculas y borrará filas vacías.")

archivo_subido = st.file_uploader("Selecciona tu archivo Excel", type=["xlsx"])

if archivo_subido is not None:
    try:
        df_procesado = cargar_y_procesar_datos(archivo_subido)
        
        # Identificar columnas clave dinámicamente
        col_medico = 'MEDICO' if 'MEDICO' in df_procesado.columns else df_procesado.columns[0]
        col_genero = next((c for c in df_procesado.columns if c in ['GENERO', 'SEXO']), None)
        
        # Filtros laterales interactivos
        st.sidebar.header("Filtros del Tablero")
        medicos_sel = st.sidebar.multiselect("Filtrar por Médico", options=df_procesado[col_medico].dropna().unique())
        
        if medicos_sel:
            df_procesado = df_procesado[df_procesado[col_medico].isin(medicos_sel)]
            
        # 4. KPIs Principales
        st.markdown("### 📈 Indicadores de Rendimiento")
        c1, c2, c3 = st.columns(3)
        
        total_visitas = len(df_procesado)
        total_efectivas = df_procesado['PRODUCTIVIDAD'].sum()
        rendimiento = (total_efectivas / total_visitas) * 100 if total_visitas > 0 else 0
        
        c1.metric("Visitas Asignadas Reales", total_visitas)
        c2.metric("Consultas Efectivas", total_efectivas)
        c3.metric("Porcentaje de Efectividad", f"{rendimiento:.1f}%")
        
        st.markdown("---")
        
        # 5. Gráficos de Productividad y Metas
        colA, colB = st.columns(2)
        
        with colA:
            st.subheader("Productividad Total vs Meta (162 Consultas)")
            df_medicos = df_procesado.groupby(col_medico)['PRODUCTIVIDAD'].sum().reset_index()
            df_medicos = df_medicos.sort_values(by='PRODUCTIVIDAD', ascending=False)
            
            fig_bar = px.bar(
                df_medicos, x=col_medico, y='PRODUCTIVIDAD', text='PRODUCTIVIDAD',
                color='PRODUCTIVIDAD', color_continuous_scale='Greens'
            )
            fig_bar.add_hline(y=162, line_dash="dash", line_color="red", annotation_text="Meta: 162", annotation_position="top left")
            st.plotly_chart(fig_bar, use_container_width=True)
            
        with colB:
            st.subheader("Evolución de Consultas")
            if 'FECHA' in df_procesado.columns:
                df_procesado['FECHA_FORMATO'] = df_procesado['FECHA'].dt.strftime('%Y-%m-%d')
                df_linea = df_procesado.groupby(['FECHA_FORMATO', col_medico])['PRODUCTIVIDAD'].sum().reset_index()
                df_linea = df_linea.sort_values('FECHA_FORMATO')
                
                fig_linea = px.line(
                    df_linea, x='FECHA_FORMATO', y='PRODUCTIVIDAD', color=col_medico,
                    title="Comparativa Diaria", markers=True
                )
                st.plotly_chart(fig_linea, use_container_width=True)
            else:
                st.warning("No se encontró la columna 'FECHA' para hacer la gráfica temporal.")
            
        # 6. Gráficos Secundarios
        st.markdown("---")
        colC, colD = st.columns(2)
        
        with colC:
            st.subheader("Motivos de Visita")
            fig_pie = px.pie(df_procesado, names='ESTADO_VISITA', hole=0.4, color_discrete_sequence=['#2ecc71', '#e67e22', '#e74c3c'])
            st.plotly_chart(fig_pie, use_container_width=True)
            
        with colD:
            st.subheader("Análisis por Género")
            if col_genero:
                fig_gen = px.pie(df_procesado, names=col_genero, hole=0.4, color_discrete_sequence=['#3498db', '#e74c3c', '#9b59b6'])
                st.plotly_chart(fig_gen, use_container_width=True)
            else:
                st.info("Para ver el gráfico de género, incluye una columna 'GENERO' o 'SEXO'.")
                
        # 7. Vista de datos final
        st.markdown("### 📋 Datos Analizados y Limpios")
        st.dataframe(df_procesado, use_container_width=True)

    except Exception as e:
        st.error(f"Error en la estructura del archivo. Verifica que corresponda al formato correcto. Detalle: {e}")