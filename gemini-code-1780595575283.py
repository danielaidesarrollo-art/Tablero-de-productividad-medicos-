import streamlit as st
import pandas as pd
import openpyxl
import plotly.express as px

# 1. Configuración de la página
st.set_page_config(page_title="Tablero Virrey Solís - Productividad", layout="wide")

# --- COLORES INSTITUCIONALES VIRREY SOLIS ---
COLOR_AZUL = "#004B87"
COLOR_VERDE = "#00A650"
COLOR_ALERTA = "#E74C3C" 

# --- DICCIONARIO DE CORRECCIÓN DE NOMBRES ---
# Aquí puedes agregar errores de tipeo para que el sistema los unifique automáticamente
CORRECCIONES_MEDICOS = {
    "KARON": "CAROL",
    "KAROL": "CAROL",
    "JHON ERIC": "JHON ERIK",
}

# --- FUNCIONES PARA DETECTAR COLORES EN EXCEL ---
def obtener_color_hex(cell):
    if not hasattr(cell, 'fill') or not cell.fill or not cell.fill.start_color:
        return ""
    return str(cell.fill.start_color.index).upper()

def es_color_cafe(color_str):
    if color_str in ['00000000', '000000', 'NONE', '']: 
        return False
    # Evitar amarillos
    if 'FF' in color_str and any(y in color_str for y in ['FFFF00', 'FFF2CC', 'FFFF99', 'FFFFCC', 'FFEB9C', 'FFFFC000']):
        return False
    # Evitar azules (usados para títulos)
    if es_color_azul(color_str): 
        return False
    return True

def es_color_azul(color_str):
    # Hexadecimales de azules comunes en Excel
    azules = ['FF00B0F0', 'FF0070C0', 'FF4472C4', 'FF5B9BD5', 'FF1F497D', 'FFDEEBF7', 'FFBDD7EE', 'FF9BC2E6']
    return any(azul in color_str for azul in azules)

# 2. PROCESAMIENTO Y LIMPIEZA DE DATOS
@st.cache_data
def cargar_y_procesar_datos(archivo_subido):
    # Lectura con pandas y limpieza de columnas
    df = pd.read_excel(archivo_subido)
    df.columns = df.columns.astype(str).str.strip().str.upper()
    
    # Rebobinar para lectura de colores
    archivo_subido.seek(0)
    wb = openpyxl.load_workbook(archivo_subido, data_only=True)
    hoja = wb.active
    
    estados = []
    filas_validas = []
    max_filas = len(df) + 1
    
    # Escaneo de colores fila por fila
    for idx, row in enumerate(hoja.iter_rows(min_row=2, max_row=max_filas, max_col=hoja.max_column)):
        color_primera_celda = obtener_color_hex(row[0])
        
        # Ignorar filas de títulos (Azules)
        if es_color_azul(color_primera_celda):
            filas_validas.append(False)
            estados.append("Ignorar")
            continue
            
        filas_validas.append(True)
        
        # Evaluar la columna F (Índice 5: Documento)
        cell_f = row[5] if len(row) > 5 else None 
        color_f = obtener_color_hex(cell_f)
        
        f_es_cafe = es_color_cafe(color_f)
        total_cafe_fila = sum(1 for cell in row if es_color_cafe(obtener_color_hex(cell)))
        
        # Lógica de cancelación
        if f_es_cafe:
            if total_cafe_fila > 3: 
                estados.append("Cancelada (Por el paciente)")
            else:
                estados.append("Cancelada (Médico no alcanzó)")
        else:
            estados.append("Consulta Efectiva")
            
    # Nivelar listas en caso de desfase
    if len(estados) > len(df): estados = estados[:len(df)]
    elif len(estados) < len(df): estados.extend(["Consulta Efectiva"] * (len(df) - len(estados)))
    
    if len(filas_validas) > len(df): filas_validas = filas_validas[:len(df)]
    elif len(filas_validas) < len(df): filas_validas.extend([True] * (len(df) - len(filas_validas)))
            
    df['ESTADO_VISITA'] = estados
    df['ES_VALIDA'] = filas_validas
    
    # Eliminar filas azules (títulos)
    df = df[df['ES_VALIDA'] == True].copy()
    
    # Sumar productividad
    df['PRODUCTIVIDAD'] = df['ESTADO_VISITA'].apply(lambda x: 1 if x == "Consulta Efectiva" else 0)
    
    # Limpieza de textos (mayúsculas y quitar nulos fantasma)
    columnas_texto = df.select_dtypes(include=['object']).columns
    for col in columnas_texto:
        df[col] = df[col].astype(str).str.strip().str.upper()
        df[col] = df[col].replace({'NAN': None, 'NONE': None, '': None, 'NAT': None})
        
    # Limpieza de vacíos reales y unificación de nombres de médicos
    col_medico = 'MEDICO' if 'MEDICO' in df.columns else df.columns[0]
    if col_medico in df.columns:
        df[col_medico] = df[col_medico].replace(CORRECCIONES_MEDICOS)
        df.dropna(subset=[col_medico], inplace=True)
        
    # Formatear fechas si existe la columna
    if 'FECHA' in df.columns:
        df['FECHA'] = pd.to_datetime(df['FECHA'], errors='coerce')
        
    return df

# 3. INTERFAZ DEL TABLERO
# Buscará el archivo "logo.png" guardado en tu repositorio de GitHub
st.image("logo.png", width=250)
st.title("📊 Tablero de Control y Productividad")
st.markdown("**Gestión de Atención Domiciliaria**")

archivo_subido = st.file_uploader("Carga la programación diaria en Excel (.xlsx)", type=["xlsx"])

if archivo_subido is not None:
    try:
        df = cargar_y_procesar_datos(archivo_subido)
        col_medico = 'MEDICO' if 'MEDICO' in df.columns else df.columns[0]
        
        # Filtros laterales
        st.sidebar.header("Filtros de Análisis")
        medicos_sel = st.sidebar.multiselect("Filtrar por Médico", options=df[col_medico].dropna().unique())
        if medicos_sel: 
            df = df[df[col_medico].isin(medicos_sel)]
            
        # 4. KPIs Principales
        st.markdown("### 📈 Indicadores Globales")
        c1, c2, c3 = st.columns(3)
        total_visitas = len(df)
        total_efectivas = df['PRODUCTIVIDAD'].sum()
        rendimiento = (total_efectivas / total_visitas) * 100 if total_visitas > 0 else 0
        
        c1.metric("Visitas Asignadas", total_visitas)
        c2.metric("Consultas Efectivas", total_efectivas)
        c3.metric("Efectividad", f"{rendimiento:.1f}%")
        st.markdown("---")
        
        # 5. Gráficos de Productividad
        colA, colB = st.columns(2)
        
        with colA:
            st.subheader("Productividad por Médico vs Meta (162)")
            df_med = df.groupby(col_medico)['PRODUCTIVIDAD'].sum().reset_index()
            fig_bar = px.bar(
                df_med.sort_values('PRODUCTIVIDAD', ascending=False), 
                x=col_medico, y='PRODUCTIVIDAD', text='PRODUCTIVIDAD',
                color_discrete_sequence=[COLOR_VERDE]
            )
            fig_bar.add_hline(y=162, line_dash="dash", line_color=COLOR_AZUL, annotation_text="Meta 162")
            st.plotly_chart(fig_bar, use_container_width=True)
            
        with colB:
            st.subheader("Evolución de Consultas Efectivas")
            if 'FECHA' in df.columns:
                df['FECHA_FORMATO'] = df['FECHA'].dt.strftime('%Y-%m-%d')
                df_linea = df.groupby(['FECHA_FORMATO', col_medico])['PRODUCTIVIDAD'].sum().reset_index()
                fig_linea = px.line(
                    df_linea.sort_values('FECHA_FORMATO'), 
                    x='FECHA_FORMATO', y='PRODUCTIVIDAD', color=col_medico,
                    markers=True
                )
                st.plotly_chart(fig_linea, use_container_width=True)
            else:
                st.warning("Agrega una columna 'FECHA' para ver el análisis temporal.")
        
        # 6. Análisis de Cancelaciones
        st.markdown("---")
        st.markdown(f"<h3 style='color: {COLOR_AZUL};'>🚨 Análisis de Cancelaciones</h3>", unsafe_allow_html=True)
        
        df_canceladas = df[df['ESTADO_VISITA'].str.contains("Cancelada", na=False)]
        
        if not df_canceladas.empty:
            colC, colD = st.columns(2)
            
            with colC:
                st.subheader("⏱️ Jornada con más cancelaciones")
                col_jornada = next((c for c in df.columns if 'JORNADA' in c), None)
                if col_jornada:
                    df_jornada = df_canceladas.groupby(col_jornada).size().reset_index(name='Cantidad')
                    fig_jornada = px.bar(
                        df_jornada.sort_values('Cantidad', ascending=False), 
                        x=col_jornada, y='Cantidad', text='Cantidad',
                        color_discrete_sequence=[COLOR_ALERTA]
                    )
                    fig_jornada.update_layout(xaxis_title="Jornada", yaxis_title="Total Cancelaciones")
                    st.plotly_chart(fig_jornada, use_container_width=True)
                else:
                    st.info("Agrega una columna 'JORNADA' a tu Excel para activar este gráfico.")
            
            with colD:
                st.subheader("📍 Zonas con mayores cancelaciones")
                col_lat = next((c for c in df.columns if 'LATITUD' in c or 'LAT' in c), None)
                col_lon = next((c for c in df.columns if 'LONGITUD' in c or 'LON' in c), None)
                col_zona = next((c for c in df.columns if c in ['ZONA', 'LOCALIDAD', 'BARRIO']), None)
                
                if col_lat and col_lon:
                    st.map(df_canceladas, latitude=col_lat, longitude=col_lon, color=COLOR_ALERTA)
                elif col_zona:
                    df_zonas = df_canceladas.groupby(col_zona).size().reset_index(name='Cancelaciones')
                    fig_zona = px.bar(df_zonas.sort_values('Cancelaciones'), x='Cancelaciones', y=col_zona, orientation='h', color_discrete_sequence=[COLOR_AZUL])
                    st.plotly_chart(fig_zona, use_container_width=True)
                else:
                    st.info("Agrega columnas de 'ZONA' o 'LOCALIDAD' para activar el análisis geográfico.")
        else:
            st.success("¡Excelente! No se registran cancelaciones en este periodo.")
            
        # 7. Tabla Detallada
        st.markdown("---")
        st.markdown("### 📋 Datos Analizados y Limpios")
        columnas_finales = [c for c in df.columns if c not in ['ES_VALIDA', 'FECHA_FORMATO']]
        st.dataframe(df[columnas_finales], use_container_width=True)

    except Exception as e:
        st.error(f"Error procesando el archivo: {e}. Revisa que las columnas coincidan con el formato esperado.")